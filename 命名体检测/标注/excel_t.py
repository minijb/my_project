from openpyxl import load_workbook
wb = load_workbook(r"C:\Users\14143\Desktop\研究生\命名体识别\标注.xlsx")
ws = wb["Sheet1"]
print(ws['A1'].value)
wb.save(r"C:\Users\14143\Desktop\研究生\命名体识别\test.xlsx")
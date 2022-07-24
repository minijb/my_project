#coding:utf-8
from openpyxl import load_workbook


with open("base.txt",'r',encoding='utf-8') as f:
    file_name  = f.readline().strip()
    nn = f.readline().strip()

nn_list = nn.split("，")
print(nn_list)


sheet = "Sheet2"
row_start = 24
column_start = 2

def print_ex(sheet,row_start,column_start,nn_list):
    wb = load_workbook(file_name)
    wb.save(file_name+".dump")
    ws = wb[sheet]
    for i,nn in enumerate(nn_list):
        ws.cell(row=row_start,column=column_start+i).value=nn
    wb.save(file_name)

print_ex(sheet,row_start,column_start,nn_list)